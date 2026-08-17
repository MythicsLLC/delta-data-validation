"""
validation_core.py
===================
Self-contained data-comparison engine — a faithful, offline port of the
``_run_validation_job`` Polars pipeline in ``Server/Main.py``.

This module has ZERO dependency on FastAPI, the job-store, or the network.
It is used by the standalone Tkinter app (``post_validation_app.py``) so the
whole tool can be frozen into a single .exe with PyInstaller.

Every cleaning / normalization / diff rule below is intentionally kept
byte-for-byte identical to Main.py's pipeline, INCLUDING the fix applied to
``_pl_clean_num_expr`` (blank numeric cells must cast to Polars ``null``, not
the string "NaN" -> float NaN, because Polars' float comparison operators
treat NaN as greater than any value, e.g. ``NaN > 0.0001`` is True. Mapping
blanks to NaN made two blank cells on both sides of a row register as a
false-positive discrepancy.)
"""

from __future__ import annotations

import os
import re
import csv
import io
import json
import gc
import time
import shutil
import threading
import concurrent.futures as cf
from datetime import datetime
from typing import Dict, List, Optional, Callable

import pandas as pd
import polars as pl

import xlsxwriter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================================
# Cancellation
# ============================================================================

class ValidationCancelled(Exception):
    """Raised when the user cancels a running validation from the GUI."""


# ============================================================================
# Constants
# ============================================================================

EXCEL_MAX_ROWS = 1_048_000          # hard Excel row limit minus headroom
_EXCEL_ROW_CAP = 1_000_000          # rows written to the xlsx tabs; beyond this -> CSV

_COMMON_DATE_FORMATS = [
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%d",
    "%Y/%m/%d",
    "%d-%m-%Y",
    "%d/%m/%Y",
    "%m/%d/%Y",
    "%m-%d-%Y",
    "%d-%m-%Y %H:%M:%S",
    "%d/%m/%Y %H:%M:%S",
    "%m/%d/%Y %H:%M:%S",
]

# Compiled once at module load — fast date-like regex (no dateutil)
_DATE_LIKE_RE = re.compile(
    r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}'       # YYYY-MM-DD, YYYY/MM/DD
    r'|^\d{1,2}[-/]\d{1,2}[-/]\d{4}'       # DD-MM-YYYY, MM/DD/YYYY
    r'|^\d{1,2}[-/]\d{1,2}[-/]\d{2}\b'     # DD-MM-YY
    r'|^\d{4}\d{2}\d{2}$'                   # YYYYMMDD
    r'|^\d{4}[-/]\d{1,2}[-/]\d{1,2}[T ]',  # ISO-8601 with time
    re.ASCII,
)

INTERNAL_KEY = "_derived_key"


# ============================================================================
# File reading helpers  (ported from Main.py)
# ============================================================================

def list_sheets(path: str) -> List[str]:
    """Return worksheet names for an Excel file; [] for CSV (no sheets)."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return []
    xls = pd.ExcelFile(path, engine="openpyxl")
    return xls.sheet_names


def read_columns_only(path: str, sheet: Optional[str] = None) -> List[str]:
    """Extract only the header row — fast, skips all data rows."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as f:
            first_line = f.readline()
        cols = [c.strip() for c in next(csv.reader([first_line])) if c.strip()]
        if cols:
            return cols
    else:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[sheet] if (sheet and sheet in wb.sheetnames) else wb.active
            cols = [str(c.value).strip() for c in next(ws.iter_rows(max_row=1)) if c.value is not None]
            wb.close()
            if cols:
                return cols
        except Exception:
            pass
    df = pd.read_excel(path, sheet_name=sheet or 0, nrows=0)
    return list(df.columns.astype(str))


def _polars_read_file(file_path: str, sheet_name=None):
    """Read CSV/Excel into a Polars LazyFrame. Mirrors Main.py's _polars_read_file."""
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".csv":
        return pl.scan_csv(
            file_path, infer_schema_length=0,
            try_parse_dates=False, null_values=[""],
            truncate_ragged_lines=True
        )

    sheet_id_param = None
    sheet_name_param = None
    if sheet_name is None:
        sheet_id_param = 1
    elif isinstance(sheet_name, int):
        sheet_id_param = max(sheet_name, 1)
    elif isinstance(sheet_name, str):
        stripped = sheet_name.strip()
        if not stripped:
            sheet_id_param = 1
        else:
            try:
                idx = int(stripped)
                sheet_id_param = max(idx + 1, 1) if idx == 0 else idx
            except ValueError:
                sheet_name_param = stripped

    df = None
    _sheet_kw: dict = {}
    if sheet_id_param is not None:
        _sheet_kw["sheet_id"] = sheet_id_param
    elif sheet_name_param is not None:
        _sheet_kw["sheet_name"] = sheet_name_param

    try:
        df = pl.read_excel(file_path, engine="calamine", **_sheet_kw)
    except Exception:
        df = None

    if df is None:
        sp = 0
        if sheet_name_param:
            sp = sheet_name_param
        elif sheet_id_param:
            sp = max(sheet_id_param - 1, 0)
        pdf = pd.read_excel(file_path, sheet_name=sp, dtype=str, engine="openpyxl")
        df = pl.from_pandas(pdf)
        del pdf

    _non_utf8 = {c: pl.Utf8 for c in df.columns if df.schema[c] != pl.Utf8}
    if _non_utf8:
        df = df.cast(_non_utf8)
    renames = {c: c.strip() for c in df.columns if c != c.strip()}
    if renames:
        df = df.rename(renames)
    return df.lazy()


# ============================================================================
# Cleaning / normalization expressions  (ported from Main.py — bug fixed)
# ============================================================================

def _pl_clean_str_expr(col_name, case_sensitive=True):
    expr = (
        pl.col(col_name).cast(pl.Utf8).fill_null("")
        .str.replace_all(r"[\p{Cf}\p{Cc}]", "")
        .str.replace_all(r"[\s\p{Z}]+", " ")
        .str.strip_chars()
        .str.replace_all(",", "")
        .str.replace_all(r"\.0+$", "")
        .str.replace_all("(?i)^nan$", "")
        .str.replace_all("(?i)^none$", "")
        .str.strip_chars()
        .fill_null("")
    )
    if not case_sensitive:
        expr = expr.str.to_lowercase()
    return expr


def _pl_clean_num_expr(col_name):
    """Clean & cast to Float64 for numeric comparison.

    Blanks are left as "" so they cast to a true Polars null (NOT the float
    NaN sentinel) — see module docstring for why this matters.
    """
    return (
        pl.col(col_name).cast(pl.Utf8).fill_null("")
        .str.replace_all(r"[\p{Cf}\p{Cc}]", "")
        .str.replace_all(r"[\s\p{Z}]+", " ")
        .str.strip_chars()
        .str.replace_all("%", "")
        .str.replace_all(",", "")
        .str.strip_chars()
        .cast(pl.Float64, strict=False)
        .round(4)
    )


def _pl_normalize_key_cols(df, key_cols):
    """Normalize key columns: strip invisible chars, collapse whitespace,
    remove .0 suffix, blank sentinels (nan/None/NaT)."""
    exprs = []
    for col in key_cols:
        if col not in df.columns:
            continue
        dtype = df.schema[col]
        if dtype in (pl.Float32, pl.Float64):
            _as_int = pl.col(col).cast(pl.Int64, strict=False)
            _initial = pl.when(
                _as_int.is_not_null() & (_as_int.cast(pl.Float64) == pl.col(col))
            ).then(_as_int.cast(pl.Utf8)).otherwise(pl.col(col).cast(pl.Utf8))
        else:
            _initial = pl.col(col).cast(pl.Utf8)
        expr = (
            _initial.fill_null("")
            .str.replace_all(r"[\p{Cf}\p{Cc}]", "")
            .str.replace_all(r"[\s\p{Z}]+", " ")
            .str.strip_chars()
            .str.replace(r"\.0+$", "", literal=False)
            .str.replace_all(r"^(?:nan|None|NaN|NaT)$", "")
        )
        exprs.append(expr.alias(col))
    return df.with_columns(exprs) if exprs else df


def _pl_gen_composite_key(key_cols):
    return pl.concat_str(
        [pl.col(c).cast(pl.Utf8).fill_null("").str.replace_all(r"\s+", "").str.to_lowercase() for c in key_cols],
        separator="|",
    )


def _pl_add_positional_key(df, key_col):
    """Append row-position-within-group suffix to key_col to handle duplicate keys safely."""
    return (
        df.with_columns(
            pl.int_range(pl.len()).over(key_col).cast(pl.Utf8).alias("_rn")
        )
        .with_columns(
            (pl.col(key_col) + pl.lit("|_rn=") + pl.col("_rn")).alias(key_col)
        )
        .drop("_rn")
    )


def _detect_date_fmt(values, formats=_COMMON_DATE_FORMATS):
    vals = [str(v).strip() for v in values[:30]
            if v is not None and str(v).strip() not in ("", "nan", "None", "NaN", "NaT")]
    if not vals:
        return None
    for fmt in formats:
        ok = 0
        for v in vals[:15]:
            try:
                datetime.strptime(v, fmt)
                ok += 1
            except Exception:
                pass
        if ok >= max(len(vals[:15]) * 0.6, 1):
            return fmt
    return None


def _pl_apply_date_normalization(df, date_cols, all_cols, auto_detect=False, compare_cols=None):
    """Batched date normalization — single pass, Polars-native str.to_date fallback."""
    _SENTINELS = {'', 'nan', 'None', 'NaN', 'NaT'}
    actual_date_cols = set(date_cols) if date_cols else set()

    if auto_detect and compare_cols:
        candidate_cols = [c for c in compare_cols if c not in actual_date_cols and c in df.columns]
        if candidate_cols:
            sample_df = df.select(
                [pl.col(c).cast(pl.Utf8).fill_null("") for c in candidate_cols]
            ).head(50)
            for col in candidate_cols:
                vals = [v.strip() for v in sample_df[col].to_list() if v.strip() not in _SENTINELS]
                if vals and sum(1 for v in vals if _DATE_LIKE_RE.match(v)) >= max(len(vals) * 0.6, 1):
                    actual_date_cols.add(col)

    valid_date_cols = [c for c in actual_date_cols if c in all_cols and c in df.columns]
    if not valid_date_cols:
        return df

    format_sample = df.select(
        [pl.col(c).cast(pl.Utf8).fill_null("") for c in valid_date_cols]
    ).head(50)

    date_exprs = []
    for col in valid_date_cols:
        sample = [v.strip() for v in format_sample[col].to_list() if v.strip() not in _SENTINELS]
        fmt = _detect_date_fmt(sample)
        if fmt:
            if "H" in fmt or "M" in fmt:
                expr = (
                    pl.col(col).cast(pl.Utf8).fill_null("")
                    .str.strptime(pl.Datetime, fmt, strict=False)
                    .dt.strftime("%Y/%m/%d").fill_null("")
                )
            else:
                expr = (
                    pl.col(col).cast(pl.Utf8).fill_null("")
                    .str.strptime(pl.Date, fmt, strict=False)
                    .dt.strftime("%Y/%m/%d").fill_null("")
                )
        else:
            expr = (
                pl.col(col).cast(pl.Utf8).fill_null("")
                .str.to_date(format=None, strict=False)
                .dt.strftime("%Y/%m/%d")
                .fill_null(pl.col(col).cast(pl.Utf8).fill_null(""))
            )
        date_exprs.append(expr.alias(col))

    if date_exprs:
        try:
            df = df.with_columns(date_exprs)
        except Exception:
            for single_expr in date_exprs:
                try:
                    df = df.with_columns([single_expr])
                except Exception:
                    pass
    return df


def _pl_detect_numeric_cols(df_l, df_o, cols_to_compare):
    """Vectorized numeric-column detection — one combined select() across all columns."""
    numeric_columns = set()
    sentinel = ["", "nan", "None", "NaN"]

    valid_cols = [c for c in cols_to_compare if c in df_l.columns and c in df_o.columns]
    if not valid_cols:
        return numeric_columns

    l_sample = df_l.select([pl.col(c).cast(pl.Utf8).alias(c) for c in valid_cols]).head(2500)
    o_sample = df_o.select([pl.col(c).cast(pl.Utf8).alias(c) for c in valid_cols]).head(2500)
    combined = pl.concat([l_sample, o_sample])

    n = len(valid_cols)
    try:
        row = combined.select(
            [
                (
                    ~pl.col(c).is_in(sentinel) &
                    pl.col(c).str.replace_all("[%,]", "").cast(pl.Float64, strict=False).is_not_null()
                ).sum().alias(f"_nc_{i}")
                for i, c in enumerate(valid_cols)
            ] +
            [
                (~pl.col(c).is_in(sentinel)).sum().alias(f"_tot_{i}")
                for i, c in enumerate(valid_cols)
            ] +
            [
                (~pl.col(c).is_in(sentinel) & pl.col(c).str.contains(r"\-")).sum().alias(f"_dc_{i}")
                for i, c in enumerate(valid_cols)
            ]
        ).row(0)
    except Exception:
        return numeric_columns

    for i, col in enumerate(valid_cols):
        tot = row[n + i]
        if tot == 0:
            continue
        nc = row[i]
        dc = row[2 * n + i]
        if (nc / tot) >= 0.8 and (dc / tot) < 0.2:
            numeric_columns.add(col)

    return numeric_columns


def _safe_sheet_name(name: str, max_len: int = 31) -> str:
    for ch in r'\/*?:[]':
        name = name.replace(ch, '_')
    return name[:max_len]


def write_df_excel_paginated(writer, df: pd.DataFrame, base_sheet_name: str, max_rows: int = EXCEL_MAX_ROWS):
    """Ported from Main.py's write_df_excel_paginated — used for the full-data sheet append."""
    safe_base = _safe_sheet_name(base_sheet_name, max_len=28)

    if len(df) <= max_rows:
        df.to_excel(writer, index=False, sheet_name=safe_base)
        return [safe_base]

    sheet_names = []
    for idx, start in enumerate(range(0, len(df), max_rows), start=1):
        sheet_name = f"{safe_base}_{idx}"
        df.iloc[start:start + max_rows].to_excel(writer, index=False, sheet_name=sheet_name)
        sheet_names.append(sheet_name)

    return sheet_names


# ============================================================================
# Main pipeline  (ported from Main.py's _run_validation_job)
# ============================================================================

def run_validation(
    p: dict,
    progress_cb: Optional[Callable[[int, str], None]] = None,
    cancel_event: Optional[threading.Event] = None,
) -> dict:
    """
    Run the full validation pipeline synchronously and write the report to
    ``p["output_dir"]``.

    Required keys in ``p``:
      legacy_path, oracle_path, legacy_filename, oracle_filename,
      mappings (dict: source_col -> target_col), keyColumns (list),
      output_dir
    Optional keys (defaults shown):
      includedColumns=[], dateColumns=[], timestampColumns=[],
      dateColumnstarget=[], timestampColumnstarget=[],
      legacySheet=None, oracleSheet=None, includeSourceTargetFiles=False,
      sourceLabel="Source", targetLabel="Target", caseSensitive=True

    Returns a dict with output_path + summary counts. Raises on error,
    raises ValidationCancelled if cancel_event was set.
    """

    def _progress(pct, stage):
        if progress_cb:
            progress_cb(pct, stage)

    def _check_cancelled():
        if cancel_event is not None and cancel_event.is_set():
            raise ValidationCancelled("Validation cancelled by user")

    output_dir = p["output_dir"]
    os.makedirs(output_dir, exist_ok=True)

    src_label = (p.get("sourceLabel") or "Source").strip() or "Source"
    tgt_label = (p.get("targetLabel") or "Target").strip() or "Target"
    case_sensitive = p.get("caseSensitive", True)

    report_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    main_output_path = os.path.join(output_dir, f"DataValidationResults_{report_ts}.xlsx")
    full_data_csv_path = os.path.join(output_dir, f"FullData_{report_ts}.csv")
    discrepancies_csv_path = os.path.join(output_dir, f"DataDiscrepancies_Full_{report_ts}.csv")
    missing_oracle_csv_path = os.path.join(output_dir, f"Missing_In_{tgt_label}_{report_ts}.csv")
    missing_ps_csv_path = os.path.join(output_dir, f"Missing_In_{src_label}_{report_ts}.csv")

    start_time = time.time()

    # ── Stage 1: Parse Inputs ──────────────────────────────────────────
    _progress(2, "Parsing configuration")
    mappings_dict: Dict[str, str] = p["mappings"]
    included_cols_list: List[str] = p.get("includedColumns", [])
    key_cols_list: List[str] = p["keyColumns"]
    legacy_date_set = set(p.get("dateColumns", []) + p.get("timestampColumns", []))
    target_date_set = set(p.get("dateColumnstarget", []) + p.get("timestampColumnstarget", []))

    if not mappings_dict:
        raise ValueError("Mappings dictionary is empty")
    if not key_cols_list:
        raise ValueError("Key columns list is empty")

    config_rows = []
    for source_col, target_col in mappings_dict.items():
        config_rows.append({
            f"{src_label} Column": source_col,
            f"{tgt_label} Column": target_col,
            "Is Key?": "Yes" if source_col in key_cols_list else "No",
            "Is Date?": "Yes" if (source_col in legacy_date_set or target_col in target_date_set) else "No",
            "Validate": "Yes",
            "Include in Report": "Yes" if source_col in included_cols_list else "No",
        })
    config_df = pd.DataFrame(config_rows)

    _check_cancelled()
    # ── Stage 2: Load Files — parallel, stay in Polars ─────────────────
    _progress(8, "Reading source and target files (parallel)")

    def _load_frame(path, sheet):
        return _polars_read_file(path, sheet).collect()

    with cf.ThreadPoolExecutor(max_workers=2, thread_name_prefix="file-load") as _load_pool:
        _f_legacy = _load_pool.submit(_load_frame, p["legacy_path"], p.get("legacySheet"))
        _f_oracle = _load_pool.submit(_load_frame, p["oracle_path"], p.get("oracleSheet"))
        pl_legacy = _f_legacy.result()
        pl_oracle = _f_oracle.result()

    legacy_count, oracle_count = len(pl_legacy), len(pl_oracle)
    _progress(20, f"Files loaded — {legacy_count:,} + {oracle_count:,} rows")

    include_src_tgt = p.get("includeSourceTargetFiles", False)

    # ── Stage 3: Column Alignment ──────────────────────────────────────
    _progress(22, "Aligning column names")

    oracle_to_legacy = {v: k for k, v in mappings_dict.items()}
    oracle_to_legacy_lower = {v.strip().lower(): k for k, v in mappings_dict.items()}

    cols_to_rename_pl: Dict[str, str] = {}
    for col in pl_oracle.columns:
        if col in oracle_to_legacy:
            cols_to_rename_pl[col] = oracle_to_legacy[col]
        elif col.strip().lower() in oracle_to_legacy_lower:
            cols_to_rename_pl[col] = oracle_to_legacy_lower[col.strip().lower()]
    if cols_to_rename_pl:
        pl_oracle = pl_oracle.rename(cols_to_rename_pl)

    legacy_mapped_lower = {k.strip().lower(): k for k in mappings_dict.keys()}
    legacy_rename_pl: Dict[str, str] = {}
    for col in pl_legacy.columns:
        lk = col.strip().lower()
        if lk in legacy_mapped_lower and col != legacy_mapped_lower[lk]:
            legacy_rename_pl[col] = legacy_mapped_lower[lk]
    if legacy_rename_pl:
        pl_legacy = pl_legacy.rename(legacy_rename_pl)

    legacy_cols = pl_legacy.columns
    oracle_cols = pl_oracle.columns

    oracle_col_lower = {c.strip().lower(): c for c in oracle_cols}
    legacy_col_lower = {c.strip().lower(): c for c in legacy_cols}
    key_cols_list = list(dict.fromkeys(
        oracle_col_lower.get(k.strip().lower(), k) for k in key_cols_list
    ))

    resolved_mappings: Dict[str, str] = {}
    for l_col, o_col in mappings_dict.items():
        actual_legacy = legacy_col_lower.get(l_col.strip().lower(), l_col)
        actual_oracle = oracle_col_lower.get(l_col.strip().lower(), l_col)
        common_name = actual_oracle if actual_oracle in oracle_cols else actual_legacy
        resolved_mappings[common_name] = o_col
    mappings_dict = resolved_mappings

    included_cols_list = list(dict.fromkeys(
        oracle_col_lower.get(c.strip().lower(), legacy_col_lower.get(c.strip().lower(), c))
        for c in included_cols_list
    ))
    legacy_date_set = {legacy_col_lower.get(c.strip().lower(), c) for c in legacy_date_set}
    target_date_set_resolved = set()
    for c in target_date_set:
        mapped = oracle_to_legacy.get(c, c)
        target_date_set_resolved.add(oracle_col_lower.get(mapped.strip().lower(), mapped))
    target_date_set = target_date_set_resolved

    missing_keys = [k for k in key_cols_list if k not in pl_oracle.columns]
    if missing_keys:
        raise ValueError(f"Key columns missing in target after rename: {missing_keys}")

    col_errors = []
    for l_col, o_col in mappings_dict.items():
        if l_col not in pl_legacy.columns:
            col_errors.append(f"{src_label} column '{l_col}' not found")
        if l_col not in pl_oracle.columns:
            col_errors.append(f"{tgt_label} column '{o_col}' not found after rename")
    if col_errors:
        raise ValueError(f"Column errors: {'; '.join(col_errors)}")

    cols_to_compare = [k for k in mappings_dict if k in pl_legacy.columns and k in pl_oracle.columns]
    num_comparison_cols = len(cols_to_compare)
    _progress(25, f"Validating {num_comparison_cols} mapped columns")

    _check_cancelled()
    # ── Stage 4: Normalize keys & dates ────────────────────────────────
    _progress(28, "Normalising keys & dates")

    pl_legacy = _pl_normalize_key_cols(pl_legacy, key_cols_list)
    pl_oracle = _pl_normalize_key_cols(pl_oracle, key_cols_list)

    legacy_date_cols_explicit = {c for c in legacy_date_set if c in pl_legacy.columns}
    target_date_cols_explicit = set()
    for col in target_date_set:
        mapped = oracle_to_legacy.get(col, col)
        if mapped in pl_oracle.columns:
            target_date_cols_explicit.add(mapped)

    with cf.ThreadPoolExecutor(max_workers=2, thread_name_prefix="date-norm") as _dn_pool:
        _f_l_dn = _dn_pool.submit(
            _pl_apply_date_normalization,
            pl_legacy, legacy_date_cols_explicit, list(pl_legacy.columns),
            True, cols_to_compare
        )
        _f_o_dn = _dn_pool.submit(
            _pl_apply_date_normalization,
            pl_oracle, target_date_cols_explicit, list(pl_oracle.columns),
            True, cols_to_compare
        )
        pl_legacy = _f_l_dn.result()
        pl_oracle = _f_o_dn.result()

    pl_legacy = _pl_normalize_key_cols(pl_legacy, key_cols_list)
    pl_oracle = _pl_normalize_key_cols(pl_oracle, key_cols_list)

    _progress(32, "Generating composite keys")
    _key_expr = _pl_gen_composite_key(key_cols_list)
    with cf.ThreadPoolExecutor(max_workers=2, thread_name_prefix="key-gen") as _kg_pool:
        _f_l_kg = _kg_pool.submit(lambda df: df.with_columns(_key_expr.alias(INTERNAL_KEY)), pl_legacy)
        _f_o_kg = _kg_pool.submit(lambda df: df.with_columns(_key_expr.alias(INTERNAL_KEY)), pl_oracle)
        pl_legacy = _f_l_kg.result()
        pl_oracle = _f_o_kg.result()
    _progress(35, "Keys generated — starting comparison")

    # ── Stage 5: Numeric detection ──────────────────────────────────────
    _progress(37, "Detecting column data types")
    numeric_cols = _pl_detect_numeric_cols(pl_legacy, pl_oracle, cols_to_compare)

    # ── Stage 5b: Duplicate key handling ────────────────────────────────
    _progress(38, "Handling duplicate keys")
    _ik = INTERNAL_KEY
    with cf.ThreadPoolExecutor(max_workers=2, thread_name_prefix="dup-check") as _dup_pool:
        _f_l_dup = _dup_pool.submit(lambda df: df.select(pl.col(_ik).is_duplicated().sum()).item(), pl_legacy)
        _f_o_dup = _dup_pool.submit(lambda df: df.select(pl.col(_ik).is_duplicated().sum()).item(), pl_oracle)
        legacy_key_dupes = _f_l_dup.result()
        oracle_key_dupes = _f_o_dup.result()

    pl_legacy_base = pl_legacy
    pl_oracle_base = pl_oracle

    if legacy_key_dupes > 0 or oracle_key_dupes > 0:
        with cf.ThreadPoolExecutor(max_workers=2, thread_name_prefix="pos-key") as _pk_pool:
            _f_l_pk = _pk_pool.submit(_pl_add_positional_key, pl_legacy, INTERNAL_KEY)
            _f_o_pk = _pk_pool.submit(_pl_add_positional_key, pl_oracle, INTERNAL_KEY)
            pl_legacy = _f_l_pk.result()
            pl_oracle = _f_o_pk.result()

    # ── Stage 6: Polars join ────────────────────────────────────────────
    _progress(40, "Joining source & target")

    compare_cols_needed = [INTERNAL_KEY] + cols_to_compare
    pl_l_cmp = pl_legacy.select([c for c in compare_cols_needed if c in pl_legacy.columns])
    pl_o_cmp = pl_oracle.select([c for c in compare_cols_needed if c in pl_oracle.columns])

    joined = pl_l_cmp.join(pl_o_cmp, on=INTERNAL_KEY, suffix="_T", how="inner")
    del pl_l_cmp, pl_o_cmp
    gc.collect()

    matched_count = len(joined)
    _progress(42, f"Joined {matched_count:,} matched rows — comparing")

    max_expected = max(legacy_count, oracle_count) * 3
    if matched_count > max_expected:
        raise ValueError(
            f"Join produced {matched_count:,} rows (source has {legacy_count:,}). "
            f"Key columns may not be unique. Please verify key column selection."
        )

    matched_keys_pl = joined.select(INTERNAL_KEY).unique()

    if include_src_tgt:
        _progress(44, "Writing matched data to CSV")
        full_select = (
            [pl.col(c).cast(pl.Utf8).fill_null("").alias(f"{c}_S") for c in cols_to_compare]
            + [pl.col(f"{c}_T").cast(pl.Utf8).fill_null("").alias(f"{c}_T") for c in cols_to_compare]
            + [pl.lit("MATCHED").alias("Record Status")]
        )
        joined.select(full_select).write_csv(full_data_csv_path)
    gc.collect()

    _check_cancelled()
    # ── Stage 6.2/7: Single-scan flag pass + mismatched-row extraction ──
    _progress(46, "Computing diff flags")

    diff_flag_exprs = []
    valid_compare_cols = []
    for col in cols_to_compare:
        o_col = f"{col}_T"
        if o_col not in joined.columns:
            continue
        valid_compare_cols.append(col)
        if col in numeric_cols:
            cn = _pl_clean_num_expr(col)
            co_expr = _pl_clean_num_expr(o_col)
            flag = (
                (cn.is_not_null() & co_expr.is_not_null() & ((cn - co_expr).abs() > 0.0001))
                | (cn.is_null() ^ co_expr.is_null())
            )
        else:
            cs = _pl_clean_str_expr(col, case_sensitive)
            co_expr = _pl_clean_str_expr(o_col, case_sensitive)
            flag = cs.fill_null("") != co_expr.fill_null("")
        diff_flag_exprs.append(flag.alias(f"__diff_{col}"))

    flags_df = joined.lazy().select(
        [pl.col(INTERNAL_KEY)] + diff_flag_exprs
    ).collect()

    cols_with_diffs = [c for c in valid_compare_cols if flags_df[f"__diff_{c}"].any()]

    _progress(55, "Extracting discrepancies")

    _EMPTY_DISC = {
        INTERNAL_KEY: pl.Series([], dtype=pl.Utf8),
        "Column Name": pl.Series([], dtype=pl.Utf8),
        f"{src_label} Value": pl.Series([], dtype=pl.Utf8),
        f"{tgt_label} Value": pl.Series([], dtype=pl.Utf8),
    }

    if cols_with_diffs:
        any_diff_expr = pl.col(f"__diff_{cols_with_diffs[0]}")
        for _c in cols_with_diffs[1:]:
            any_diff_expr = any_diff_expr | pl.col(f"__diff_{_c}")
        any_diff_mask = flags_df.select(any_diff_expr.alias("_any"))["_any"]

        needed_val_cols = (
            [INTERNAL_KEY]
            + [c for c in cols_with_diffs if c in joined.columns]
            + [f"{c}_T" for c in cols_with_diffs if f"{c}_T" in joined.columns]
        )
        mismatched_vals = joined.filter(any_diff_mask).select(needed_val_cols)

        flags_mismatch = flags_df.filter(any_diff_mask).select(
            [f"__diff_{c}" for c in cols_with_diffs]
        )
        work = pl.concat([mismatched_vals, flags_mismatch], how="horizontal")
        del mismatched_vals, flags_mismatch, flags_df, any_diff_mask
        gc.collect()

        disc_parts = []
        for col in cols_with_diffs:
            col_label = f"{col} - {mappings_dict.get(col, col)}"
            part = work.filter(pl.col(f"__diff_{col}")).select([
                pl.col(INTERNAL_KEY),
                pl.lit(col_label).alias("Column Name"),
                pl.col(col).cast(pl.Utf8).fill_null("").alias(f"{src_label} Value"),
                pl.col(f"{col}_T").cast(pl.Utf8).fill_null("").alias(f"{tgt_label} Value"),
            ])
            if len(part) > 0:
                disc_parts.append(part)
        del work

        discrepancies_pl = pl.concat(disc_parts) if disc_parts else pl.DataFrame(_EMPTY_DISC)
        del disc_parts
    else:
        del flags_df
        discrepancies_pl = pl.DataFrame(_EMPTY_DISC)
    gc.collect()

    context_cols_order = list(dict.fromkeys(key_cols_list + included_cols_list))
    valid_ctx = [c for c in context_cols_order if c in pl_legacy.columns and c != INTERNAL_KEY]
    if valid_ctx:
        ctx_pl = pl_legacy.select([INTERNAL_KEY] + valid_ctx).unique(subset=[INTERNAL_KEY])
        discrepancies_pl = discrepancies_pl.join(ctx_pl, on=INTERNAL_KEY, how="left")

    total_discrepancies = len(discrepancies_pl)
    unique_discrepant_records = (
        discrepancies_pl.select(INTERNAL_KEY).n_unique()
        if total_discrepancies > 0 else 0
    )

    column_discrepancy_counts: list = []
    if total_discrepancies > 0 and "Column Name" in discrepancies_pl.columns:
        _col_cnt = (
            discrepancies_pl.select("Column Name")
            .group_by("Column Name")
            .agg(pl.len().alias("cnt"))
            .sort("cnt", descending=True)
        )
        for _r in _col_cnt.iter_rows(named=True):
            column_discrepancy_counts.append(["", _r["Column Name"], int(_r["cnt"]), "", "", ""])

    if total_discrepancies > _EXCEL_ROW_CAP:
        discrepancies_pl.drop(INTERNAL_KEY).write_csv(discrepancies_csv_path)

    if total_discrepancies > 0:
        disc_for_excel = (
            discrepancies_pl.head(_EXCEL_ROW_CAP) if total_discrepancies > _EXCEL_ROW_CAP
            else discrepancies_pl
        )
        validation_df = disc_for_excel.drop(INTERNAL_KEY).to_pandas()
        del discrepancies_pl, disc_for_excel
        gc.collect()

        final_report_cols = (
            key_cols_list
            + [c for c in included_cols_list if c not in key_cols_list]
            + ["Column Name", f"{src_label} Value", f"{tgt_label} Value"]
        )
        final_report_cols = [c for c in final_report_cols if c in validation_df.columns]
        validation_df = validation_df[final_report_cols]
    else:
        del discrepancies_pl
        validation_df = pd.DataFrame([{"Status": "All mapped columns matched perfectly"}])

    _progress(62, f"Found {total_discrepancies:,} discrepancies")

    if "Column Name" in validation_df.columns and not validation_df.empty:
        sort_cols = ["Column Name"] + [c for c in key_cols_list if c in validation_df.columns]
        validation_df = validation_df.sort_values(by=sort_cols, kind="mergesort").reset_index(drop=True)

    comment_cols = ["Mythics Comments", "Oracle Comments", "ParkView Comments"]
    for col in comment_cols:
        validation_df[col] = ""

    _check_cancelled()
    # ── Stage 9: Missing Records — set-based ────────────────────────────
    _progress(64, "Finding missing records")

    _oracle_key_set = set(pl_oracle_base[INTERNAL_KEY].to_list())
    _legacy_key_set = set(pl_legacy_base[INTERNAL_KEY].to_list())
    _missing_in_oracle = _legacy_key_set - _oracle_key_set
    _missing_in_ps = _oracle_key_set - _legacy_key_set

    legacy_only_pl = pl_legacy_base.filter(pl.col(INTERNAL_KEY).is_in(_missing_in_oracle))
    oracle_only_pl = pl_oracle_base.filter(pl.col(INTERNAL_KEY).is_in(_missing_in_ps))
    del _oracle_key_set, _legacy_key_set, _missing_in_oracle, _missing_in_ps

    count_missing_oracle = len(legacy_only_pl)
    count_missing_ps = len(oracle_only_pl)
    _progress(68, f"Missing: {count_missing_oracle:,} in {tgt_label}, {count_missing_ps:,} in {src_label}")

    _MISSING_EXCEL_CAP = _EXCEL_ROW_CAP
    if count_missing_oracle > _MISSING_EXCEL_CAP:
        legacy_only_pl.drop(INTERNAL_KEY).write_csv(missing_oracle_csv_path)
    if count_missing_ps > _MISSING_EXCEL_CAP:
        oracle_only_pl.drop(INTERNAL_KEY).write_csv(missing_ps_csv_path)

    legacy_for_excel = (
        legacy_only_pl.head(_MISSING_EXCEL_CAP) if count_missing_oracle > _MISSING_EXCEL_CAP
        else legacy_only_pl
    )
    oracle_for_excel = (
        oracle_only_pl.head(_MISSING_EXCEL_CAP) if count_missing_ps > _MISSING_EXCEL_CAP
        else oracle_only_pl
    )
    _disp_l = [c for c in legacy_for_excel.columns if c != INTERNAL_KEY]
    _disp_o = [c for c in oracle_for_excel.columns if c != INTERNAL_KEY]

    def _cast_int_floats(df):
        exprs = []
        for col, dt in df.schema.items():
            if dt not in (pl.Float32, pl.Float64):
                continue
            _as_int = pl.col(col).cast(pl.Int64, strict=False)
            exprs.append(
                pl.when(_as_int.is_not_null() & (_as_int.cast(pl.Float64) == pl.col(col)))
                .then(_as_int.cast(pl.Utf8))
                .otherwise(pl.col(col).cast(pl.Utf8))
                .alias(col)
            )
        return df.with_columns(exprs) if exprs else df

    legacy_only_df = (
        _cast_int_floats(legacy_for_excel.select(_disp_l) if _disp_l else legacy_for_excel.drop(INTERNAL_KEY))
    ).to_pandas()
    oracle_only_df = (
        _cast_int_floats(oracle_for_excel.select(_disp_o) if _disp_o else oracle_for_excel.drop(INTERNAL_KEY))
    ).to_pandas()
    del legacy_for_excel, oracle_for_excel
    for col in comment_cols:
        legacy_only_df[col] = ""
        oracle_only_df[col] = ""

    # ── Stage 10: Append missing rows to full-data CSV ──────────────────
    _progress(78, "Appending missing records to full data CSV")

    if include_src_tgt and os.path.exists(full_data_csv_path):
        if count_missing_oracle > 0:
            l_s = [(pl.col(c).cast(pl.Utf8).fill_null("") if c in legacy_only_pl.columns else pl.lit("")).alias(f"{c}_S")
                   for c in cols_to_compare]
            l_t = [pl.lit("").alias(f"{c}_T") for c in cols_to_compare]
            with open(full_data_csv_path, "ab") as _f:
                legacy_only_pl.select(
                    l_s + l_t + [pl.lit("MISSING_IN_TARGET").alias("Record Status")]
                ).write_csv(_f, include_header=False)
        if count_missing_ps > 0:
            o_s = [pl.lit("").alias(f"{c}_S") for c in cols_to_compare]
            o_t = [(pl.col(c).cast(pl.Utf8).fill_null("") if c in oracle_only_pl.columns else pl.lit("")).alias(f"{c}_T")
                   for c in cols_to_compare]
            with open(full_data_csv_path, "ab") as _f:
                oracle_only_pl.select(
                    o_s + o_t + [pl.lit("MISSING_IN_SOURCE").alias("Record Status")]
                ).write_csv(_f, include_header=False)

    full_data_for_excel = None
    if include_src_tgt and os.path.exists(full_data_csv_path):
        try:
            full_data_for_excel = pd.read_csv(full_data_csv_path, dtype=str, nrows=_EXCEL_ROW_CAP)
        except Exception:
            full_data_for_excel = None

    del joined, pl_legacy, pl_oracle, matched_keys_pl, legacy_only_pl, oracle_only_pl, pl_legacy_base, pl_oracle_base
    gc.collect()

    # ── Stage 13: Summary ────────────────────────────────────────────────
    _progress(82, "Generating summary statistics")

    grand_total = count_missing_oracle + count_missing_ps + total_discrepancies
    summary_data = [
        ["", "Comparison Statistics", "", "", "", ""],
        ["", f"{src_label} File Name", p["legacy_filename"], "", "", ""],
        ["", f"{src_label} Records Count", f"{legacy_count:,}", "", "", ""],
        ["", f"{tgt_label} File Name", p["oracle_filename"], "", "", ""],
        ["", f"{tgt_label} Records Count", f"{oracle_count:,}", "", "", ""],
        ["", "Validation DateTime", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "", "", ""],
        ["", "", "", "", "", ""],
        ["", "Missing Records Summary", "", "Mythics Comments", "Oracle Comments", "ParkView Comments"],
        ["", f"Records Missing in {src_label}", f"{count_missing_ps:,}", "", "", ""],
        ["", f"Records Missing in {tgt_label}", f"{count_missing_oracle:,}", "", "", ""],
        ["", "Total Missing Records", f"{count_missing_oracle + count_missing_ps:,}", "", "", ""],
        ["", "", "", "", "", ""],
        ["", "Data Discrepancies Summary", "", "Mythics Comments", "Oracle Comments", "ParkView Comments"],
        *column_discrepancy_counts,
        ["", "Unique Records with Discrepancies", f"{unique_discrepant_records:,}", "", "", ""],
        ["", "Total Data Discrepancies (cell-level)", f"{total_discrepancies:,}", "", "", ""],
        ["", "", "", "", "", ""],
        ["", "Total Validation Issues", f"{grand_total:,}", "", "", ""],
    ]

    # ── Stage 14: Write styled Excel (xlsxwriter) ───────────────────────
    _progress(85, "Writing styled Excel report")

    sheet_missing_ps = _safe_sheet_name(f"Missing in {src_label}")
    sheet_missing_oc = _safe_sheet_name(f"Missing in {tgt_label}")
    sheet_discrepancies = "Data Discrepancies"
    sheet_full_data = _safe_sheet_name(f"{src_label} - {tgt_label} Data", max_len=28)

    font_white = Font(name="Calibri", size=8, color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    fill_header_ps = PatternFill("solid", fgColor="1F497D")
    fill_header_oc = PatternFill("solid", fgColor="31869B")
    border_thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def _style_full_data_header(ws, n_compare_cols):
        border_ps_last = Border(left=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"), right=Side(style="medium"))
        border_oc_first = Border(left=Side(style="medium"), top=Side(style="thin"), bottom=Side(style="thin"), right=Side(style="thin"))
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = font_white
            cell.alignment = align_center
            if cell.column <= n_compare_cols:
                cell.fill = fill_header_ps
                cell.border = border_ps_last if cell.column == n_compare_cols else border_thin
            else:
                cell.fill = fill_header_oc
                cell.border = border_oc_first if cell.column == n_compare_cols + 1 else border_thin
        for col_cells in ws.iter_cols(max_row=min(50, ws.max_row)):
            ml = max((len(str(c.value)) if c.value else 0) for c in col_cells)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
                max((ml + 2) * 1.2, 10), 60
            )

    _wb = xlsxwriter.Workbook(main_output_path, {'in_memory': False})
    try:
        def _fmt(**kw):
            return _wb.add_format({'font_name': 'Calibri', 'font_size': 8, **kw})

        _f_hdr_ps = _fmt(bold=True, bg_color='#1F497D', font_color='white', border=1, align='center', valign='vcenter')
        _f_hdr_oc = _fmt(bold=True, bg_color='#31869B', font_color='white', border=1, align='center', valign='vcenter')
        _f_hdr_err = _fmt(bold=True, bg_color='#C0504D', font_color='white', border=1, align='center', valign='vcenter')
        _f_hdr_grn = _fmt(bold=True, bg_color='#00B050', font_color='white', border=1, align='center', valign='vcenter')
        _f_hdr_org = _fmt(bold=True, bg_color='#FF9900', font_color='black', border=1, align='center', valign='vcenter')
        _f_bdr = _fmt(border=1)
        _fs_grn_hdr = _fmt(bold=True, bg_color='#00B050', font_color='white', border=1, align='center', valign='vcenter')
        _fs_grn_lbl = _fmt(bg_color='#00B050', font_color='white', border=1)
        _fs_val = _fmt(align='center', border=1)
        _fs_grey_l = _fmt(bold=True, bg_color='#D9D9D9', border=1)
        _fs_grey_v = _fmt(bold=True, bg_color='#D9D9D9', align='center', border=1)
        _fs_org = _fmt(bold=True, bg_color='#FF9900', font_color='black', align='center', border=1)
        _fs_bdr = _fmt(border=1)

        _comment_set = set(comment_cols)

        def _write_styled(df, sn, hdr_fmt):
            safe_sn = _safe_sheet_name(sn, max_len=28)
            if df.empty:
                ws_empty = _wb.add_worksheet(safe_sn)
                for ci, col in enumerate(df.columns):
                    ws_empty.write(0, ci, col, hdr_fmt)
                return [safe_sn]
            chunks = (
                [(1, df)] if len(df) <= EXCEL_MAX_ROWS
                else [(pgn, df.iloc[s:s + EXCEL_MAX_ROWS])
                      for pgn, s in enumerate(range(0, len(df), EXCEL_MAX_ROWS), 1)]
            )
            names = []
            for pg_num, chunk in chunks:
                pg_sn = safe_sn if pg_num == 1 else f"{safe_sn}_{pg_num}"
                ws = _wb.add_worksheet(pg_sn)
                _arr = chunk.to_numpy(na_value='')
                for _ci in range(_arr.shape[1]):
                    ws.write_column(1, _ci, _arr[:, _ci].tolist())
                comment_col_idx = set()
                for ci, col in enumerate(chunk.columns):
                    fmt = _f_hdr_org if col in _comment_set else hdr_fmt
                    ws.write(0, ci, col, fmt)
                    if col in _comment_set:
                        comment_col_idx.add(ci)
                ws.freeze_panes(1, 0)
                sample = chunk.head(50)
                for ci, col in enumerate(chunk.columns):
                    vals = sample.iloc[:, ci]
                    ml = max(
                        (len(str(v)) for v in vals if v is not None and str(v) not in ('', 'nan')),
                        default=4,
                    )
                    ml = max(ml, len(str(col)))
                    ws.set_column(ci, ci, min(max((ml + 2) * 1.2, 10), 60),
                                   _f_bdr if ci in comment_col_idx else None)
                names.append(pg_sn)
            return names

        ws_sum = _wb.add_worksheet("Summary")
        ws_sum.hide_gridlines(2)
        ws_sum.set_column('A:A', 2)
        ws_sum.set_column('B:B', 45)
        for _ch in ('C', 'D', 'E', 'F'):
            ws_sum.set_column(f'{_ch}:{_ch}', 25)

        _write_styled(oracle_only_df, sheet_missing_ps, _f_hdr_ps)
        _write_styled(legacy_only_df, sheet_missing_oc, _f_hdr_oc)
        _write_styled(validation_df, sheet_discrepancies, _f_hdr_err)

        if not config_df.empty:
            ws_cfg = _wb.add_worksheet("Configuration")
            _arr_cfg = config_df.to_numpy(na_value='')
            for ri, _row in enumerate(_arr_cfg, 1):
                for ci, val in enumerate(_row):
                    ws_cfg.write(ri, ci, val if val is not None else '')
            for ci, col in enumerate(config_df.columns):
                ws_cfg.write(0, ci, col, _f_hdr_grn)
            ws_cfg.freeze_panes(1, 0)
            for ci, col in enumerate(config_df.columns):
                ml = max((len(str(v)) for v in config_df.iloc[:, ci] if v), default=4)
                ml = max(ml, len(str(col)))
                ws_cfg.set_column(ci, ci, min(max(ml + 2, 12), 45), _f_bdr)
            ws_cfg.hide()

        for ri, row_data in enumerate(summary_data):
            row6 = (list(row_data) + ['', '', '', '', '', ''])[:6]
            _, val_b, val_c, val_d, val_e, val_f = row6
            if not val_b:
                continue
            val_b_str = str(val_b)
            if val_b in ("Missing Records Summary", "Data Discrepancies Summary"):
                ws_sum.merge_range(ri, 1, ri, 2, val_b, _fs_grn_hdr)
                ws_sum.set_row(ri, 20)
                ws_sum.write(ri, 3, val_d or '', _fs_org)
                ws_sum.write(ri, 4, val_e or '', _fs_org)
                ws_sum.write(ri, 5, val_f or '', _fs_org)
            elif 'Comparison Statistics' in val_b_str:
                ws_sum.merge_range(ri, 1, ri, 2, val_b, _fs_grn_hdr)
                ws_sum.set_row(ri, 20)
                ws_sum.write(ri, 3, '', _fs_bdr)
                ws_sum.write(ri, 4, '', _fs_bdr)
                ws_sum.write(ri, 5, '', _fs_bdr)
            elif 'Total' in val_b_str:
                ws_sum.write(ri, 1, val_b, _fs_grey_l)
                ws_sum.write(ri, 2, val_c or '', _fs_grey_v)
                ws_sum.write(ri, 3, '', _fs_bdr)
                ws_sum.write(ri, 4, '', _fs_bdr)
                ws_sum.write(ri, 5, '', _fs_bdr)
            else:
                ws_sum.write(ri, 1, val_b, _fs_grn_lbl)
                ws_sum.write(ri, 2, val_c or '', _fs_val)
                ws_sum.write(ri, 3, '', _fs_bdr)
                ws_sum.write(ri, 4, '', _fs_bdr)
                ws_sum.write(ri, 5, '', _fs_bdr)
    finally:
        _wb.close()

    _progress(92, "Excel file written")

    if include_src_tgt and full_data_for_excel is not None and not full_data_for_excel.empty:
        _progress(94, "Appending source/target data sheet")
        try:
            with pd.ExcelWriter(
                main_output_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
            ) as app_writer:
                _fd_names = write_df_excel_paginated(app_writer, full_data_for_excel, sheet_full_data)
                _wb_app = app_writer.book
                for _sn in _fd_names:
                    if _sn in _wb_app.sheetnames:
                        _style_full_data_header(_wb_app[_sn], num_comparison_cols)
        except Exception as _fd_err:
            pass  # full-data sheet is best-effort; core report still succeeds

    if full_data_for_excel is not None:
        del full_data_for_excel
        gc.collect()

    _progress(96, "Finalizing output")
    elapsed = time.time() - start_time

    _progress(100, "Validation complete")

    return {
        "output_path": main_output_path,
        "elapsed_seconds": round(elapsed, 2),
        "legacy_count": legacy_count,
        "oracle_count": oracle_count,
        "total_discrepancies": total_discrepancies,
        "unique_discrepant_records": unique_discrepant_records,
        "count_missing_in_source": count_missing_ps,
        "count_missing_in_target": count_missing_oracle,
        "discrepancies_csv": discrepancies_csv_path if total_discrepancies > _EXCEL_ROW_CAP else None,
        "missing_in_target_csv": missing_oracle_csv_path if count_missing_oracle > _MISSING_EXCEL_CAP else None,
        "missing_in_source_csv": missing_ps_csv_path if count_missing_ps > _MISSING_EXCEL_CAP else None,
        "full_data_csv": full_data_csv_path if include_src_tgt and os.path.exists(full_data_csv_path) else None,
    }
